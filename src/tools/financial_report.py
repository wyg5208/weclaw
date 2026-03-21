"""财务报告工具 — 财务报表生成与分析：资产负债表、利润表、现金流量表、财务比率分析。

支持的操作：
- 生成资产负债表（Excel格式，含专业样式）
- 生成利润表
- 生成现金流量表
- 财务比率分析（输出Markdown报告）
- 多格式导出（xlsx/pdf/csv）
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.tools.base import ActionDef, BaseTool, ToolResult, ToolResultStatus

logger = logging.getLogger(__name__)


class FinancialReportTool(BaseTool):
    """财务报告工具。

    支持财务报表生成与分析：资产负债表、利润表、现金流量表、财务比率分析。
    """

    name = "financial_report"
    emoji = "💰"
    title = "财务报告"
    description = "财务报表生成与分析工具"
    timeout = 120

    # Excel 样式定义
    HEADER_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    TITLE_FONT = Font(name="微软雅黑", size=16, bold=True)
    SECTION_FONT = Font(name="微软雅黑", size=11, bold=True)
    NORMAL_FONT = Font(name="微软雅黑", size=10)
    TOTAL_FONT = Font(name="微软雅黑", size=11, bold=True)
    TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

    def __init__(self, output_dir: str | None = None) -> None:
        """初始化财务报告工具。

        Args:
            output_dir: 输出目录，默认为项目的 generated/日期/ 目录
        """
        super().__init__()
        self.output_dir = (
            Path(output_dir)
            if output_dir
            else Path(__file__).parent.parent.parent / "generated" / datetime.now().strftime("%Y-%m-%d")
        )
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def get_actions(self) -> list[ActionDef]:
        return [
            ActionDef(
                name="generate_balance_sheet",
                description="生成资产负债表（Excel格式，含专业样式）",
                parameters={
                    "company_name": {
                        "type": "string",
                        "description": "公司名称",
                    },
                    "date": {
                        "type": "string",
                        "description": "报表日期，如 2024-12-31",
                    },
                    "assets": {
                        "type": "object",
                        "description": "资产数据，包含流动资产和非流动资产",
                        "properties": {
                            "流动资产": {
                                "type": "object",
                                "description": "流动资产明细，如 {货币资金: 1000000, 应收账款: 500000}",
                            },
                            "非流动资产": {
                                "type": "object",
                                "description": "非流动资产明细，如 {固定资产: 2000000, 无形资产: 300000}",
                            },
                        },
                    },
                    "liabilities": {
                        "type": "object",
                        "description": "负债数据，包含流动负债和非流动负债",
                        "properties": {
                            "流动负债": {
                                "type": "object",
                                "description": "流动负债明细，如 {应付账款: 300000, 短期借款: 200000}",
                            },
                            "非流动负债": {
                                "type": "object",
                                "description": "非流动负债明细，如 {长期借款: 500000}",
                            },
                        },
                    },
                    "equity": {
                        "type": "object",
                        "description": "所有者权益数据，如 {实收资本: 1000000, 资本公积: 200000, 留存收益: 500000}",
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "输出文件名（不含扩展名），可选",
                    },
                },
                required_params=["company_name", "date", "assets", "liabilities", "equity"],
            ),
            ActionDef(
                name="generate_income_statement",
                description="生成利润表（Excel格式，含专业样式）",
                parameters={
                    "company_name": {
                        "type": "string",
                        "description": "公司名称",
                    },
                    "period": {
                        "type": "string",
                        "description": "报表期间，如 2024年度 或 2024年1-6月",
                    },
                    "revenue": {
                        "type": "object",
                        "description": "收入数据",
                        "properties": {
                            "主营业务收入": {
                                "type": "number",
                                "description": "主营业务收入金额",
                            },
                            "其他业务收入": {
                                "type": "number",
                                "description": "其他业务收入金额",
                            },
                        },
                    },
                    "costs": {
                        "type": "object",
                        "description": "成本费用数据",
                        "properties": {
                            "主营业务成本": {
                                "type": "number",
                                "description": "主营业务成本金额",
                            },
                            "销售费用": {
                                "type": "number",
                                "description": "销售费用金额",
                            },
                            "管理费用": {
                                "type": "number",
                                "description": "管理费用金额",
                            },
                            "财务费用": {
                                "type": "number",
                                "description": "财务费用金额",
                            },
                            "研发费用": {
                                "type": "number",
                                "description": "研发费用金额",
                            },
                        },
                    },
                    "taxes": {
                        "type": "number",
                        "description": "所得税费用",
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "输出文件名（不含扩展名），可选",
                    },
                },
                required_params=["company_name", "period", "revenue", "costs", "taxes"],
            ),
            ActionDef(
                name="generate_cash_flow",
                description="生成现金流量表（Excel格式，含专业样式）",
                parameters={
                    "company_name": {
                        "type": "string",
                        "description": "公司名称",
                    },
                    "period": {
                        "type": "string",
                        "description": "报表期间，如 2024年度",
                    },
                    "operating": {
                        "type": "object",
                        "description": "经营活动现金流数据",
                        "properties": {
                            "销售商品收到的现金": {"type": "number"},
                            "收到的税费返还": {"type": "number"},
                            "收到其他与经营活动有关的现金": {"type": "number"},
                            "购买商品支付的现金": {"type": "number"},
                            "支付给职工的现金": {"type": "number"},
                            "支付的各项税费": {"type": "number"},
                            "支付其他与经营活动有关的现金": {"type": "number"},
                        },
                    },
                    "investing": {
                        "type": "object",
                        "description": "投资活动现金流数据",
                        "properties": {
                            "收回投资收到的现金": {"type": "number"},
                            "取得投资收益收到的现金": {"type": "number"},
                            "处置固定资产收到的现金": {"type": "number"},
                            "购建固定资产支付的现金": {"type": "number"},
                            "投资支付的现金": {"type": "number"},
                        },
                    },
                    "financing": {
                        "type": "object",
                        "description": "筹资活动现金流数据",
                        "properties": {
                            "吸收投资收到的现金": {"type": "number"},
                            "取得借款收到的现金": {"type": "number"},
                            "偿还债务支付的现金": {"type": "number"},
                            "分配股利支付的现金": {"type": "number"},
                            "支付利息支付的现金": {"type": "number"},
                        },
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "输出文件名（不含扩展名），可选",
                    },
                },
                required_params=["company_name", "period", "operating", "investing", "financing"],
            ),
            ActionDef(
                name="financial_analysis",
                description="财务比率分析，计算资产负债率、流动比率、净利率、ROE、ROA等指标",
                parameters={
                    "data": {
                        "type": "object",
                        "description": "财务数据",
                        "properties": {
                            "总资产": {"type": "number", "description": "总资产金额"},
                            "总负债": {"type": "number", "description": "总负债金额"},
                            "所有者权益": {"type": "number", "description": "所有者权益金额"},
                            "流动资产": {"type": "number", "description": "流动资产金额"},
                            "流动负债": {"type": "number", "description": "流动负债金额"},
                            "净利润": {"type": "number", "description": "净利润金额"},
                            "营业收入": {"type": "number", "description": "营业收入金额"},
                            "经营活动现金流净额": {"type": "number", "description": "经营活动现金流净额"},
                            "存货": {"type": "number", "description": "存货金额，可选"},
                            "应收账款": {"type": "number", "description": "应收账款金额，可选"},
                        },
                    },
                    "company_name": {
                        "type": "string",
                        "description": "公司名称，可选",
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "输出文件名（不含扩展名），可选",
                    },
                },
                required_params=["data"],
            ),
            ActionDef(
                name="export_report",
                description="导出财务报告，支持多格式导出",
                parameters={
                    "input_file": {
                        "type": "string",
                        "description": "输入文件路径（xlsx格式）",
                    },
                    "format": {
                        "type": "string",
                        "description": "导出格式",
                        "enum": ["xlsx", "pdf", "csv"],
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "输出文件名（不含扩展名），可选",
                    },
                },
                required_params=["input_file", "format"],
            ),
        ]

    async def execute(self, action: str, params: dict[str, Any]) -> ToolResult:
        """执行财务报告动作。"""
        action_map = {
            "generate_balance_sheet": self._generate_balance_sheet,
            "generate_income_statement": self._generate_income_statement,
            "generate_cash_flow": self._generate_cash_flow,
            "financial_analysis": self._financial_analysis,
            "export_report": self._export_report,
        }

        handler = action_map.get(action)
        if not handler:
            return ToolResult(
                status=ToolResultStatus.ERROR,
                error=f"不支持的动作: {action}",
            )

        try:
            return handler(params)
        except Exception as e:
            logger.error("财务报告生成失败: %s", e, exc_info=True)
            return ToolResult(
                status=ToolResultStatus.ERROR,
                error=f"财务报告生成失败: {e}",
            )

    # ------------------------------------------------------------------
    # 辅助方法
    # ------------------------------------------------------------------

    def _get_output_path(self, prefix: str, output_filename: str | None, ext: str) -> Path:
        """生成输出文件路径。"""
        if output_filename:
            return self.output_dir / f"{output_filename}.{ext}"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.output_dir / f"{prefix}_{timestamp}.{ext}"

    def _apply_cell_style(
        self,
        ws,
        row: int,
        col: int,
        value: Any,
        font: Font | None = None,
        fill: PatternFill | None = None,
        alignment: Alignment | None = None,
        number_format: str | None = None,
    ) -> None:
        """为单元格应用样式。"""
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = self.THIN_BORDER
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if number_format:
            cell.number_format = number_format

    def _format_currency(self, value: float) -> str:
        """格式化货币显示。"""
        if value >= 100000000:
            return f"{value / 100000000:.2f}亿"
        elif value >= 10000:
            return f"{value / 10000:.2f}万"
        else:
            return f"{value:.2f}"

    def _build_result(
        self,
        output_path: Path,
        operation: str,
        extra_info: str = "",
    ) -> ToolResult:
        """构建成功的返回结果。"""
        file_size = output_path.stat().st_size if output_path.exists() else 0
        output = (
            f"✅ {operation}完成\n"
            f"📁 文件: {output_path.name}\n"
            f"📂 路径: {output_path}\n"
            f"💾 大小: {file_size} 字节"
        )
        if extra_info:
            output += f"\n{extra_info}"
        return ToolResult(
            status=ToolResultStatus.SUCCESS,
            output=output,
            data={
                "file_path": str(output_path),
                "file_name": output_path.name,
                "file_size": file_size,
            },
        )

    # ------------------------------------------------------------------
    # 报表生成方法
    # ------------------------------------------------------------------

    def _generate_balance_sheet(self, params: dict[str, Any]) -> ToolResult:
        """生成资产负债表。"""
        company_name = params["company_name"]
        date = params["date"]
        assets = params["assets"]
        liabilities = params["liabilities"]
        equity = params["equity"]
        output_filename = params.get("output_filename")

        wb = Workbook()
        ws = wb.active
        ws.title = "资产负债表"

        # 设置列宽
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 18

        # 标题
        ws.merge_cells("A1:D1")
        self._apply_cell_style(ws, 1, 1, f"{company_name}", self.TITLE_FONT, alignment=self.CENTER_ALIGN)
        ws.merge_cells("A2:D2")
        self._apply_cell_style(ws, 2, 1, "资产负债表", self.TITLE_FONT, alignment=self.CENTER_ALIGN)
        ws.merge_cells("A3:D3")
        self._apply_cell_style(ws, 3, 1, f"编制日期：{date}", self.NORMAL_FONT, alignment=self.CENTER_ALIGN)
        ws.merge_cells("A4:D4")
        self._apply_cell_style(ws, 4, 1, "单位：元", self.NORMAL_FONT, alignment=self.RIGHT_ALIGN)

        # 表头
        row = 5
        self._apply_cell_style(ws, row, 1, "资产", self.HEADER_FONT, self.HEADER_FILL, self.CENTER_ALIGN)
        self._apply_cell_style(ws, row, 2, "金额", self.HEADER_FONT, self.HEADER_FILL, self.CENTER_ALIGN)
        self._apply_cell_style(ws, row, 3, "负债和所有者权益", self.HEADER_FONT, self.HEADER_FILL, self.CENTER_ALIGN)
        self._apply_cell_style(ws, row, 4, "金额", self.HEADER_FONT, self.HEADER_FILL, self.CENTER_ALIGN)

        row = 6
        total_assets = 0
        total_liabilities = 0
        total_equity = 0

        # 流动资产
        current_assets = assets.get("流动资产", {})
        self._apply_cell_style(ws, row, 1, "流动资产", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)

        # 流动负债
        current_liabilities = liabilities.get("流动负债", {})
        self._apply_cell_style(ws, row, 3, "流动负债", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 4, "", self.NORMAL_FONT)
        row += 1

        # 填充流动资产和流动负债
        max_items = max(len(current_assets), len(current_liabilities))
        asset_items = list(current_assets.items())
        liability_items = list(current_liabilities.items())

        current_assets_total = sum(current_assets.values())
        current_liabilities_total = sum(current_liabilities.values())

        for i in range(max_items):
            if i < len(asset_items):
                name, value = asset_items[i]
                self._apply_cell_style(ws, row, 1, f"  {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
                self._apply_cell_style(ws, row, 2, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            else:
                self._apply_cell_style(ws, row, 1, "", self.NORMAL_FONT)
                self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)

            if i < len(liability_items):
                name, value = liability_items[i]
                self._apply_cell_style(ws, row, 3, f"  {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
                self._apply_cell_style(ws, row, 4, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            else:
                self._apply_cell_style(ws, row, 3, "", self.NORMAL_FONT)
                self._apply_cell_style(ws, row, 4, "", self.NORMAL_FONT)
            row += 1

        # 流动资产/负债合计
        self._apply_cell_style(ws, row, 1, "流动资产合计", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, current_assets_total, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        self._apply_cell_style(ws, row, 3, "流动负债合计", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 4, current_liabilities_total, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        row += 1

        # 非流动资产
        non_current_assets = assets.get("非流动资产", {})
        self._apply_cell_style(ws, row, 1, "非流动资产", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)

        # 非流动负债
        non_current_liabilities = liabilities.get("非流动负债", {})
        self._apply_cell_style(ws, row, 3, "非流动负债", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 4, "", self.NORMAL_FONT)
        row += 1

        # 填充非流动资产和非流动负债
        max_items = max(len(non_current_assets), len(non_current_liabilities))
        asset_items = list(non_current_assets.items())
        liability_items = list(non_current_liabilities.items())

        non_current_assets_total = sum(non_current_assets.values())
        non_current_liabilities_total = sum(non_current_liabilities.values())

        for i in range(max_items):
            if i < len(asset_items):
                name, value = asset_items[i]
                self._apply_cell_style(ws, row, 1, f"  {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
                self._apply_cell_style(ws, row, 2, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            else:
                self._apply_cell_style(ws, row, 1, "", self.NORMAL_FONT)
                self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)

            if i < len(liability_items):
                name, value = liability_items[i]
                self._apply_cell_style(ws, row, 3, f"  {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
                self._apply_cell_style(ws, row, 4, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            else:
                self._apply_cell_style(ws, row, 3, "", self.NORMAL_FONT)
                self._apply_cell_style(ws, row, 4, "", self.NORMAL_FONT)
            row += 1

        # 非流动资产/负债合计
        self._apply_cell_style(ws, row, 1, "非流动资产合计", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, non_current_assets_total, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        self._apply_cell_style(ws, row, 3, "非流动负债合计", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 4, non_current_liabilities_total, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        row += 1

        # 负债合计
        total_liabilities = current_liabilities_total + non_current_liabilities_total
        self._apply_cell_style(ws, row, 1, "", self.NORMAL_FONT)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)
        self._apply_cell_style(ws, row, 3, "负债合计", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 4, total_liabilities, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        row += 1

        # 所有者权益
        self._apply_cell_style(ws, row, 1, "", self.NORMAL_FONT)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)
        self._apply_cell_style(ws, row, 3, "所有者权益", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 4, "", self.NORMAL_FONT)
        row += 1

        for name, value in equity.items():
            self._apply_cell_style(ws, row, 1, "", self.NORMAL_FONT)
            self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)
            self._apply_cell_style(ws, row, 3, f"  {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
            self._apply_cell_style(ws, row, 4, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            row += 1

        total_equity = sum(equity.values())
        self._apply_cell_style(ws, row, 1, "", self.NORMAL_FONT)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)
        self._apply_cell_style(ws, row, 3, "所有者权益合计", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 4, total_equity, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        row += 1

        # 资产合计 / 负债+权益合计
        total_assets = current_assets_total + non_current_assets_total
        total_liab_equity = total_liabilities + total_equity

        heavy_border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )
        final_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        ws.cell(row=row, column=1, value="资产总计")
        ws.cell(row=row, column=1).font = self.TOTAL_FONT
        ws.cell(row=row, column=1).fill = final_fill
        ws.cell(row=row, column=1).border = heavy_border
        ws.cell(row=row, column=1).alignment = self.LEFT_ALIGN

        ws.cell(row=row, column=2, value=total_assets)
        ws.cell(row=row, column=2).font = self.TOTAL_FONT
        ws.cell(row=row, column=2).fill = final_fill
        ws.cell(row=row, column=2).border = heavy_border
        ws.cell(row=row, column=2).alignment = self.RIGHT_ALIGN
        ws.cell(row=row, column=2).number_format = "#,##0.00"

        ws.cell(row=row, column=3, value="负债和所有者权益总计")
        ws.cell(row=row, column=3).font = self.TOTAL_FONT
        ws.cell(row=row, column=3).fill = final_fill
        ws.cell(row=row, column=3).border = heavy_border
        ws.cell(row=row, column=3).alignment = self.LEFT_ALIGN

        ws.cell(row=row, column=4, value=total_liab_equity)
        ws.cell(row=row, column=4).font = self.TOTAL_FONT
        ws.cell(row=row, column=4).fill = final_fill
        ws.cell(row=row, column=4).border = heavy_border
        ws.cell(row=row, column=4).alignment = self.RIGHT_ALIGN
        ws.cell(row=row, column=4).number_format = "#,##0.00"

        # 保存文件
        output_path = self._get_output_path("资产负债表", output_filename, "xlsx")
        wb.save(output_path)

        extra_info = (
            f"📊 资产总计: {self._format_currency(total_assets)}\n"
            f"📊 负债总计: {self._format_currency(total_liabilities)}\n"
            f"📊 所有者权益: {self._format_currency(total_equity)}"
        )
        return self._build_result(output_path, "资产负债表生成", extra_info)

    def _generate_income_statement(self, params: dict[str, Any]) -> ToolResult:
        """生成利润表。"""
        company_name = params["company_name"]
        period = params["period"]
        revenue = params["revenue"]
        costs = params["costs"]
        taxes = params["taxes"]
        output_filename = params.get("output_filename")

        wb = Workbook()
        ws = wb.active
        ws.title = "利润表"

        # 设置列宽
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20

        # 标题
        ws.merge_cells("A1:B1")
        self._apply_cell_style(ws, 1, 1, f"{company_name}", self.TITLE_FONT, alignment=self.CENTER_ALIGN)
        ws.merge_cells("A2:B2")
        self._apply_cell_style(ws, 2, 1, "利润表", self.TITLE_FONT, alignment=self.CENTER_ALIGN)
        ws.merge_cells("A3:B3")
        self._apply_cell_style(ws, 3, 1, f"报告期间：{period}", self.NORMAL_FONT, alignment=self.CENTER_ALIGN)
        ws.merge_cells("A4:B4")
        self._apply_cell_style(ws, 4, 1, "单位：元", self.NORMAL_FONT, alignment=self.RIGHT_ALIGN)

        # 表头
        row = 5
        self._apply_cell_style(ws, row, 1, "项目", self.HEADER_FONT, self.HEADER_FILL, self.CENTER_ALIGN)
        self._apply_cell_style(ws, row, 2, "本期金额", self.HEADER_FONT, self.HEADER_FILL, self.CENTER_ALIGN)
        row += 1

        # 营业收入
        self._apply_cell_style(ws, row, 1, "一、营业收入", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)
        row += 1

        total_revenue = 0
        for name, value in revenue.items():
            self._apply_cell_style(ws, row, 1, f"    {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
            self._apply_cell_style(ws, row, 2, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            total_revenue += value
            row += 1

        self._apply_cell_style(ws, row, 1, "营业收入合计", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, total_revenue, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        row += 1

        # 营业成本
        self._apply_cell_style(ws, row, 1, "二、营业成本", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)
        row += 1

        total_costs = 0
        for name, value in costs.items():
            self._apply_cell_style(ws, row, 1, f"    {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
            self._apply_cell_style(ws, row, 2, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            total_costs += value
            row += 1

        self._apply_cell_style(ws, row, 1, "营业成本合计", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, total_costs, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        row += 1

        # 营业利润
        operating_profit = total_revenue - total_costs
        self._apply_cell_style(ws, row, 1, "三、营业利润", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, operating_profit, self.SECTION_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
        row += 1

        # 利润总额（假设没有营业外收支）
        self._apply_cell_style(ws, row, 1, "四、利润总额", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, operating_profit, self.SECTION_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
        row += 1

        # 所得税费用
        self._apply_cell_style(ws, row, 1, "    减：所得税费用", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, taxes, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
        row += 1

        # 净利润
        net_profit = operating_profit - taxes
        heavy_border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )
        final_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        ws.cell(row=row, column=1, value="五、净利润")
        ws.cell(row=row, column=1).font = self.TOTAL_FONT
        ws.cell(row=row, column=1).fill = final_fill
        ws.cell(row=row, column=1).border = heavy_border
        ws.cell(row=row, column=1).alignment = self.LEFT_ALIGN

        ws.cell(row=row, column=2, value=net_profit)
        ws.cell(row=row, column=2).font = self.TOTAL_FONT
        ws.cell(row=row, column=2).fill = final_fill
        ws.cell(row=row, column=2).border = heavy_border
        ws.cell(row=row, column=2).alignment = self.RIGHT_ALIGN
        ws.cell(row=row, column=2).number_format = "#,##0.00"

        # 保存文件
        output_path = self._get_output_path("利润表", output_filename, "xlsx")
        wb.save(output_path)

        extra_info = (
            f"📊 营业收入: {self._format_currency(total_revenue)}\n"
            f"📊 营业成本: {self._format_currency(total_costs)}\n"
            f"📊 净利润: {self._format_currency(net_profit)}"
        )
        return self._build_result(output_path, "利润表生成", extra_info)

    def _generate_cash_flow(self, params: dict[str, Any]) -> ToolResult:
        """生成现金流量表。"""
        company_name = params["company_name"]
        period = params["period"]
        operating = params["operating"]
        investing = params["investing"]
        financing = params["financing"]
        output_filename = params.get("output_filename")

        wb = Workbook()
        ws = wb.active
        ws.title = "现金流量表"

        # 设置列宽
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20

        # 标题
        ws.merge_cells("A1:B1")
        self._apply_cell_style(ws, 1, 1, f"{company_name}", self.TITLE_FONT, alignment=self.CENTER_ALIGN)
        ws.merge_cells("A2:B2")
        self._apply_cell_style(ws, 2, 1, "现金流量表", self.TITLE_FONT, alignment=self.CENTER_ALIGN)
        ws.merge_cells("A3:B3")
        self._apply_cell_style(ws, 3, 1, f"报告期间：{period}", self.NORMAL_FONT, alignment=self.CENTER_ALIGN)
        ws.merge_cells("A4:B4")
        self._apply_cell_style(ws, 4, 1, "单位：元", self.NORMAL_FONT, alignment=self.RIGHT_ALIGN)

        # 表头
        row = 5
        self._apply_cell_style(ws, row, 1, "项目", self.HEADER_FONT, self.HEADER_FILL, self.CENTER_ALIGN)
        self._apply_cell_style(ws, row, 2, "本期金额", self.HEADER_FONT, self.HEADER_FILL, self.CENTER_ALIGN)
        row += 1

        # 经营活动现金流
        self._apply_cell_style(ws, row, 1, "一、经营活动产生的现金流量", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)
        row += 1

        operating_inflow = 0
        operating_outflow = 0
        inflow_items = ["销售商品收到的现金", "收到的税费返还", "收到其他与经营活动有关的现金"]

        for name, value in operating.items():
            self._apply_cell_style(ws, row, 1, f"    {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
            self._apply_cell_style(ws, row, 2, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            if any(item in name for item in inflow_items) or "收到" in name:
                operating_inflow += value
            else:
                operating_outflow += value
            row += 1

        operating_net = operating_inflow - operating_outflow
        self._apply_cell_style(ws, row, 1, "经营活动产生的现金流量净额", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, operating_net, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        row += 1

        # 投资活动现金流
        self._apply_cell_style(ws, row, 1, "二、投资活动产生的现金流量", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)
        row += 1

        investing_inflow = 0
        investing_outflow = 0
        invest_inflow_items = ["收回投资", "取得投资收益", "处置固定资产"]

        for name, value in investing.items():
            self._apply_cell_style(ws, row, 1, f"    {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
            self._apply_cell_style(ws, row, 2, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            if any(item in name for item in invest_inflow_items) or "收到" in name:
                investing_inflow += value
            else:
                investing_outflow += value
            row += 1

        investing_net = investing_inflow - investing_outflow
        self._apply_cell_style(ws, row, 1, "投资活动产生的现金流量净额", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, investing_net, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        row += 1

        # 筹资活动现金流
        self._apply_cell_style(ws, row, 1, "三、筹资活动产生的现金流量", self.SECTION_FONT, alignment=self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, "", self.NORMAL_FONT)
        row += 1

        financing_inflow = 0
        financing_outflow = 0
        finance_inflow_items = ["吸收投资", "取得借款"]

        for name, value in financing.items():
            self._apply_cell_style(ws, row, 1, f"    {name}", self.NORMAL_FONT, alignment=self.LEFT_ALIGN)
            self._apply_cell_style(ws, row, 2, value, self.NORMAL_FONT, alignment=self.RIGHT_ALIGN, number_format="#,##0.00")
            if any(item in name for item in finance_inflow_items) or "收到" in name:
                financing_inflow += value
            else:
                financing_outflow += value
            row += 1

        financing_net = financing_inflow - financing_outflow
        self._apply_cell_style(ws, row, 1, "筹资活动产生的现金流量净额", self.TOTAL_FONT, self.TOTAL_FILL, self.LEFT_ALIGN)
        self._apply_cell_style(ws, row, 2, financing_net, self.TOTAL_FONT, self.TOTAL_FILL, self.RIGHT_ALIGN, "#,##0.00")
        row += 1

        # 现金净增加额
        total_net = operating_net + investing_net + financing_net
        heavy_border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )
        final_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        ws.cell(row=row, column=1, value="四、现金及现金等价物净增加额")
        ws.cell(row=row, column=1).font = self.TOTAL_FONT
        ws.cell(row=row, column=1).fill = final_fill
        ws.cell(row=row, column=1).border = heavy_border
        ws.cell(row=row, column=1).alignment = self.LEFT_ALIGN

        ws.cell(row=row, column=2, value=total_net)
        ws.cell(row=row, column=2).font = self.TOTAL_FONT
        ws.cell(row=row, column=2).fill = final_fill
        ws.cell(row=row, column=2).border = heavy_border
        ws.cell(row=row, column=2).alignment = self.RIGHT_ALIGN
        ws.cell(row=row, column=2).number_format = "#,##0.00"

        # 保存文件
        output_path = self._get_output_path("现金流量表", output_filename, "xlsx")
        wb.save(output_path)

        extra_info = (
            f"📊 经营活动净额: {self._format_currency(operating_net)}\n"
            f"📊 投资活动净额: {self._format_currency(investing_net)}\n"
            f"📊 筹资活动净额: {self._format_currency(financing_net)}\n"
            f"📊 现金净增加额: {self._format_currency(total_net)}"
        )
        return self._build_result(output_path, "现金流量表生成", extra_info)

    def _financial_analysis(self, params: dict[str, Any]) -> ToolResult:
        """财务比率分析。"""
        data = params["data"]
        company_name = params.get("company_name", "公司")
        output_filename = params.get("output_filename")

        # 获取数据
        total_assets = data.get("总资产", 0)
        total_liabilities = data.get("总负债", 0)
        equity = data.get("所有者权益", total_assets - total_liabilities)
        current_assets = data.get("流动资产", 0)
        current_liabilities = data.get("流动负债", 0)
        net_profit = data.get("净利润", 0)
        revenue = data.get("营业收入", 0)
        operating_cash_flow = data.get("经营活动现金流净额", 0)
        inventory = data.get("存货", 0)
        receivables = data.get("应收账款", 0)

        # 计算财务比率
        ratios = {}

        # 偿债能力指标
        if total_assets > 0:
            ratios["资产负债率"] = (total_liabilities / total_assets) * 100
        if current_liabilities > 0:
            ratios["流动比率"] = current_assets / current_liabilities
            quick_assets = current_assets - inventory
            ratios["速动比率"] = quick_assets / current_liabilities

        # 盈利能力指标
        if revenue > 0:
            ratios["净利率"] = (net_profit / revenue) * 100
        if equity > 0:
            ratios["ROE (净资产收益率)"] = (net_profit / equity) * 100
        if total_assets > 0:
            ratios["ROA (总资产收益率)"] = (net_profit / total_assets) * 100

        # 运营能力指标
        if revenue > 0 and total_assets > 0:
            ratios["总资产周转率"] = revenue / total_assets
        if revenue > 0 and inventory > 0:
            ratios["存货周转率"] = revenue / inventory
        if revenue > 0 and receivables > 0:
            ratios["应收账款周转率"] = revenue / receivables

        # 现金流指标
        if net_profit != 0:
            ratios["经营现金流/净利润"] = operating_cash_flow / net_profit

        # 生成Markdown报告
        report = f"# {company_name} 财务分析报告\n\n"
        report += f"**生成日期**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"

        report += "## 一、基础数据\n\n"
        report += "| 项目 | 金额（元） |\n"
        report += "|------|----------|\n"
        report += f"| 总资产 | {total_assets:,.2f} |\n"
        report += f"| 总负债 | {total_liabilities:,.2f} |\n"
        report += f"| 所有者权益 | {equity:,.2f} |\n"
        report += f"| 流动资产 | {current_assets:,.2f} |\n"
        report += f"| 流动负债 | {current_liabilities:,.2f} |\n"
        report += f"| 净利润 | {net_profit:,.2f} |\n"
        report += f"| 营业收入 | {revenue:,.2f} |\n\n"

        report += "## 二、财务比率分析\n\n"

        report += "### 2.1 偿债能力指标\n\n"
        report += "| 指标 | 数值 | 评价 |\n"
        report += "|------|------|------|\n"
        if "资产负债率" in ratios:
            debt_ratio = ratios["资产负债率"]
            evaluation = "偏高" if debt_ratio > 70 else ("适中" if debt_ratio > 40 else "较低")
            report += f"| 资产负债率 | {debt_ratio:.2f}% | {evaluation} |\n"
        if "流动比率" in ratios:
            current_ratio = ratios["流动比率"]
            evaluation = "良好" if 1.5 <= current_ratio <= 2.5 else ("偏低" if current_ratio < 1.5 else "偏高")
            report += f"| 流动比率 | {current_ratio:.2f} | {evaluation} |\n"
        if "速动比率" in ratios:
            quick_ratio = ratios["速动比率"]
            evaluation = "良好" if quick_ratio >= 1 else "偏低"
            report += f"| 速动比率 | {quick_ratio:.2f} | {evaluation} |\n"
        report += "\n"

        report += "### 2.2 盈利能力指标\n\n"
        report += "| 指标 | 数值 | 评价 |\n"
        report += "|------|------|------|\n"
        if "净利率" in ratios:
            net_margin = ratios["净利率"]
            evaluation = "优秀" if net_margin > 15 else ("良好" if net_margin > 5 else "一般")
            report += f"| 净利率 | {net_margin:.2f}% | {evaluation} |\n"
        if "ROE (净资产收益率)" in ratios:
            roe = ratios["ROE (净资产收益率)"]
            evaluation = "优秀" if roe > 15 else ("良好" if roe > 8 else "一般")
            report += f"| ROE | {roe:.2f}% | {evaluation} |\n"
        if "ROA (总资产收益率)" in ratios:
            roa = ratios["ROA (总资产收益率)"]
            evaluation = "优秀" if roa > 10 else ("良好" if roa > 5 else "一般")
            report += f"| ROA | {roa:.2f}% | {evaluation} |\n"
        report += "\n"

        report += "### 2.3 运营能力指标\n\n"
        report += "| 指标 | 数值 |\n"
        report += "|------|------|\n"
        if "总资产周转率" in ratios:
            report += f"| 总资产周转率 | {ratios['总资产周转率']:.2f}次 |\n"
        if "存货周转率" in ratios:
            report += f"| 存货周转率 | {ratios['存货周转率']:.2f}次 |\n"
        if "应收账款周转率" in ratios:
            report += f"| 应收账款周转率 | {ratios['应收账款周转率']:.2f}次 |\n"
        report += "\n"

        if "经营现金流/净利润" in ratios:
            report += "### 2.4 现金流指标\n\n"
            report += "| 指标 | 数值 | 评价 |\n"
            report += "|------|------|------|\n"
            cash_ratio = ratios["经营现金流/净利润"]
            evaluation = "良好" if cash_ratio > 1 else "需关注"
            report += f"| 经营现金流/净利润 | {cash_ratio:.2f} | {evaluation} |\n\n"

        report += "## 三、综合评价\n\n"
        # 简单的综合评价
        good_points = []
        bad_points = []

        if ratios.get("资产负债率", 100) < 60:
            good_points.append("资产负债率处于合理水平，财务风险可控")
        else:
            bad_points.append("资产负债率偏高，需注意财务风险")

        if ratios.get("流动比率", 0) >= 1.5:
            good_points.append("流动比率良好，短期偿债能力较强")
        elif ratios.get("流动比率", 0) > 0:
            bad_points.append("流动比率偏低，短期偿债能力需加强")

        if ratios.get("净利率", 0) > 10:
            good_points.append("净利率较高，盈利能力较强")
        elif ratios.get("净利率", 0) > 0:
            good_points.append("公司保持盈利")
        else:
            bad_points.append("公司处于亏损状态")

        if good_points:
            report += "**优势**:\n"
            for point in good_points:
                report += f"- {point}\n"
            report += "\n"

        if bad_points:
            report += "**需关注**:\n"
            for point in bad_points:
                report += f"- {point}\n"
            report += "\n"

        report += "---\n*本报告由财务分析工具自动生成，仅供参考。*\n"

        # 保存报告
        output_path = self._get_output_path("财务分析报告", output_filename, "md")
        output_path.write_text(report, encoding="utf-8")

        return ToolResult(
            status=ToolResultStatus.SUCCESS,
            output=f"✅ 财务分析报告生成完成\n📁 文件: {output_path.name}\n📂 路径: {output_path}\n\n{report[:500]}...",
            data={
                "file_path": str(output_path),
                "file_name": output_path.name,
                "ratios": ratios,
            },
        )

    def _export_report(self, params: dict[str, Any]) -> ToolResult:
        """导出财务报告。"""
        input_file = params["input_file"]
        export_format = params["format"]
        output_filename = params.get("output_filename")

        input_path = Path(input_file)
        if not input_path.exists():
            return ToolResult(
                status=ToolResultStatus.ERROR,
                error=f"输入文件不存在: {input_file}",
            )

        # 读取Excel文件
        if input_path.suffix.lower() in (".xlsx", ".xls"):
            df = pd.read_excel(input_path)
        else:
            return ToolResult(
                status=ToolResultStatus.ERROR,
                error=f"不支持的输入文件格式: {input_path.suffix}",
            )

        # 生成输出路径
        output_path = self._get_output_path(input_path.stem, output_filename, export_format)

        if export_format == "xlsx":
            # 直接复制Excel文件，保留格式
            import shutil
            shutil.copy(input_path, output_path)
        elif export_format == "csv":
            df.to_csv(output_path, index=False, encoding="utf-8-sig")
        elif export_format == "pdf":
            # PDF导出需要额外的依赖，这里提供基本实现
            try:
                # 尝试使用 matplotlib 生成表格PDF
                import matplotlib.pyplot as plt
                from matplotlib.backends.backend_pdf import PdfPages

                with PdfPages(str(output_path)) as pdf:
                    fig, ax = plt.subplots(figsize=(12, len(df) * 0.4 + 2))
                    ax.axis("tight")
                    ax.axis("off")
                    
                    # 设置中文字体
                    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
                    plt.rcParams["axes.unicode_minus"] = False
                    
                    table = ax.table(
                        cellText=df.values,
                        colLabels=df.columns,
                        cellLoc="center",
                        loc="center",
                    )
                    table.auto_set_font_size(False)
                    table.set_fontsize(9)
                    table.scale(1.2, 1.5)
                    
                    pdf.savefig(fig, bbox_inches="tight")
                    plt.close()
            except Exception as e:
                return ToolResult(
                    status=ToolResultStatus.ERROR,
                    error=f"PDF导出失败: {e}。建议先导出为xlsx或csv格式。",
                )
        else:
            return ToolResult(
                status=ToolResultStatus.ERROR,
                error=f"不支持的导出格式: {export_format}",
            )

        format_map = {"xlsx": "Excel", "csv": "CSV", "pdf": "PDF"}
        extra_info = f"📤 导出格式: {format_map.get(export_format, export_format)}"
        return self._build_result(output_path, "财务报告导出", extra_info)


# 用于测试
if __name__ == "__main__":
    import asyncio

    async def test():
        tool = FinancialReportTool()
        print("Actions:", [a.name for a in tool.get_actions()])
        print("Tool registered successfully!")

        # 测试资产负债表生成
        result = await tool.execute(
            "generate_balance_sheet",
            {
                "company_name": "测试科技有限公司",
                "date": "2024-12-31",
                "assets": {
                    "流动资产": {"货币资金": 1000000, "应收账款": 500000, "存货": 300000},
                    "非流动资产": {"固定资产": 2000000, "无形资产": 300000},
                },
                "liabilities": {
                    "流动负债": {"应付账款": 300000, "短期借款": 200000},
                    "非流动负债": {"长期借款": 500000},
                },
                "equity": {"实收资本": 2000000, "资本公积": 200000, "留存收益": 600000},
            },
        )
        print("资产负债表:", result.output)

    asyncio.run(test())
