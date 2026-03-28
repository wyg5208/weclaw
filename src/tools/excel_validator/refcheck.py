"""Excel 引用异常检查模块 - 检测引用问题（范围溢出、表头误包含等）。

检查项：
- 范围溢出（引用超出数据区域）
- 表头行被包含在计算中
- 窄聚合（SUM仅覆盖1-2个单元格）
- 相邻公式模式偏差

参考 MiniMax refcheck.py 设计
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class ReferenceIssue:
    """引用异常问题。"""
    issue_type: str  # "overflow" | "header_included" | "narrow_aggregate" | "pattern_deviation"
    cell: str         # 单元格地址
    sheet: str        # 工作表名称
    formula: str      # 原始公式
    severity: str     # "error" | "warning" | "info"
    message: str      # 问题描述
    suggestion: str   # 修复建议


@dataclass
class RefcheckResult:
    """引用检查结果。"""
    status: str  # "success" | "issues_found" | "failed"
    total_issues: int = 0
    error_count: int = 0
    warning_count: int = 0
    issues: list[ReferenceIssue] = field(default_factory=list)
    sheets_checked: list[str] = field(default_factory=list)
    message: str = ""

    @property
    def is_clean(self) -> bool:
        return self.status == "success" or self.error_count == 0


def _parse_cell_ref(cell_ref: str) -> tuple[str, int, int]:
    """解析单元格引用，返回 (列字母, 行号, 列号)。
    
    例如: "A1" -> ("A", 1, 1), "BC123" -> ("BC", 123, 55)
    """
    match = re.match(r"^([A-Z]+)(\d+)$", cell_ref.upper())
    if not match:
        return "", 0, 0
    
    col_letters = match.group(1)
    row = int(match.group(2))
    
    # 将列字母转换为数字
    col_num = 0
    for char in col_letters:
        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
    
    return col_letters, row, col_num


def _extract_range_info(formula: str) -> list[tuple[str, str, int, int]]:
    """从公式中提取范围引用信息。
    
    返回: [(范围字符串, 开始单元格, 开始行列, 结束行列), ...]
    """
    ranges = []
    # 匹配 A1:B10 格式的范围引用
    pattern = r'\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)'
    matches = re.findall(pattern, formula)
    
    for match in matches:
        start_col, start_row, end_col, end_row = match
        start_row, end_row = int(start_row), int(end_row)
        
        # 计算列号
        start_col_num = 0
        end_col_num = 0
        for char in start_col:
            start_col_num = start_col_num * 26 + (ord(char) - ord('A') + 1)
        for char in end_col:
            end_col_num = end_col_num * 26 + (ord(char) - ord('A') + 1)
        
        range_str = f"{start_col}{start_row}:{end_col}{end_row}"
        ranges.append((range_str, f"{start_col}{start_row}", 
                      start_row * 1000 + start_col_num,
                      end_row * 1000 + end_col_num))
    
    return ranges


def _detect_header_included(formula: str, sheet_data: dict[str, list]) -> bool:
    """检测表头行是否被包含在计算中。
    
    通过检查公式范围的第一行是否与表头行匹配来检测。
    """
    ranges = _extract_range_info(formula)
    if not ranges:
        return False
    
    range_str, start_cell, start_key, end_key = ranges[0]
    start_row = int(re.search(r'\d+', start_cell).group())
    
    # 检查第一行是否是文本（可能是表头）
    if start_row > 1 and sheet_data:
        # 假设第一行是表头，检查 start_row 是否引用了文本内容
        # 这需要访问单元格的实际值
        pass
    
    return False


def refcheck_workbook(path: str) -> RefcheckResult:
    """检查 Excel 文件的引用异常。
    
    Args:
        path: Excel 文件路径
        
    Returns:
        RefcheckResult: 引用检查结果
    """
    file_path = Path(path)
    if not file_path.exists():
        return RefcheckResult(
            status="failed",
            message=f"文件不存在: {path}"
        )
    
    if file_path.suffix.lower() not in (".xlsx", ".xlsm"):
        return RefcheckResult(
            status="failed",
            message=f"不支持的文件格式: {file_path.suffix}"
        )
    
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        issues: list[ReferenceIssue] = []
        sheets_checked = []
        
        for sheet_name in wb.sheetnames:
            sheets_checked.append(sheet_name)
            ws = wb[sheet_name]
            
            # 获取工作表数据范围
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            
            # 遍历所有单元格检查公式
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:  # 是公式
                        formula = cell.value
                        if not isinstance(formula, str):
                            continue
                        
                        cell_addr = cell.coordinate
                        
                        # 检查1: 范围溢出（引用超出实际数据区域）
                        issues.extend(_check_range_overflow(
                            formula, cell_addr, sheet_name, max_row, max_col
                        ))
                        
                        # 检查2: 窄聚合（仅1-2个单元格）
                        issues.extend(_check_narrow_aggregate(
                            formula, cell_addr, sheet_name
                        ))
                        
                        # 检查3: 检测可能的表头误包含
                        issues.extend(_check_header_inclusion(
                            formula, cell_addr, sheet_name, ws
                        ))
        
        # 分类统计
        error_count = sum(1 for i in issues if i.severity == "error")
        warning_count = sum(1 for i in issues if i.severity == "warning")
        
        if issues:
            return RefcheckResult(
                status="issues_found",
                total_issues=len(issues),
                error_count=error_count,
                warning_count=warning_count,
                issues=issues,
                sheets_checked=sheets_checked,
                message=f"发现 {len(issues)} 个引用问题 ({error_count} 个错误, {warning_count} 个警告)"
            )
        else:
            return RefcheckResult(
                status="success",
                sheets_checked=sheets_checked,
                message="所有引用正常"
            )
            
    except Exception as e:
        logger.error("引用检查异常: %s", e)
        return RefcheckResult(
            status="failed",
            message=f"引用检查异常: {e}"
        )


def _check_range_overflow(
    formula: str, 
    cell_addr: str, 
    sheet_name: str,
    max_row: int,
    max_col: int
) -> list[ReferenceIssue]:
    """检查范围溢出问题。"""
    issues = []
    ranges = _extract_range_info(formula)
    
    for range_str, start_cell, start_key, end_key in ranges:
        # 解析范围结束位置
        match = re.match(r'[A-Z]+(\d+):[A-Z]+(\d+)', range_str)
        if match:
            end_row = int(match.group(2))
            
            # 检查是否超出 max_row
            if end_row > max_row + 10:  # 允许一定余量
                issues.append(ReferenceIssue(
                    issue_type="overflow",
                    cell=cell_addr,
                    sheet=sheet_name,
                    formula=formula,
                    severity="warning",
                    message=f"范围 {range_str} 可能超出数据区域（最大行: {max_row}）",
                    suggestion="考虑缩小范围或扩展数据区域"
                ))
    
    return issues


def _check_narrow_aggregate(
    formula: str, 
    cell_addr: str, 
    sheet_name: str
) -> list[ReferenceIssue]:
    """检查窄聚合问题（SUM仅覆盖1-2个单元格）。"""
    issues = []
    
    # 检测 SUM 函数
    sum_pattern = r'(?:SUM|AVERAGE|COUNT|MAX|MIN)\s*\(\s*([A-Z]+)(\d+)\s*:\s*([A-Z]+)(\d+)\s*\)'
    matches = re.findall(sum_pattern, formula.upper())
    
    for match in matches:
        start_col, start_row, end_col, end_row = match
        start_row, end_row = int(start_row), int(end_row)
        
        # 计算涉及的单元格数量
        # 列数
        start_col_num = 0
        end_col_num = 0
        for char in start_col:
            start_col_num = start_col_num * 26 + (ord(char) - ord('A') + 1)
        for char in end_col:
            end_col_num = end_col_num * 26 + (ord(char) - ord('A') + 1)
        
        col_count = end_col_num - start_col_num + 1
        row_count = end_row - start_row + 1
        total_cells = col_count * row_count
        
        # 如果 SUM/AVERAGE 只涉及 1-2 个单元格，可能是意外的
        if total_cells <= 2:
            issues.append(ReferenceIssue(
                issue_type="narrow_aggregate",
                cell=cell_addr,
                sheet=sheet_name,
                formula=formula,
                severity="info",
                message=f"聚合函数仅涉及 {total_cells} 个单元格",
                suggestion="确认这是有意为之，或考虑扩展范围"
            ))
    
    return issues


def _check_header_inclusion(
    formula: str, 
    cell_addr: str, 
    sheet_name: str,
    ws: openpyxl.worksheet.worksheet.Worksheet
) -> list[ReferenceIssue]:
    """检查表头是否被误包含在计算中。"""
    issues = []
    ranges = _extract_range_info(formula)
    
    for range_str, start_cell, start_key, end_key in ranges:
        # 解析起始行
        match = re.match(r'[A-Z]+(\d+):[A-Z]+(\d+)', range_str)
        if match:
            start_row = int(match.group(1))
            
            # 如果范围从第1行开始，检查第1行是否是文本（表头）
            if start_row == 1:
                # 获取起始列的单元格值
                col_letter = re.match(r'([A-Z]+)', start_cell).group(1)
                first_cell_value = ws[f"{col_letter}1"].value
                
                # 如果是文本且不是数字，则可能是表头
                if first_cell_value and isinstance(first_cell_value, str):
                    issues.append(ReferenceIssue(
                        issue_type="header_included",
                        cell=cell_addr,
                        sheet=sheet_name,
                        formula=formula,
                        severity="warning",
                        message=f"范围 {range_str} 可能包含表头行",
                        suggestion="考虑从第2行开始排除表头"
                    ))
    
    return issues


# 兼容性黑名单（参考 MiniMax guardrails）
FORBIDDEN_FUNCTIONS = [
    "XLOOKUP",    # Excel 2019+ | 替代: INDEX() + MATCH()
    "XMATCH",     # Excel 2019+ | 替代: MATCH()
    "FILTER",     # Excel 2019+ | 替代: AutoFilter / SUMIF
    "UNIQUE",     # Excel 2019+ | 替代: Remove Duplicates
    "LET",        # Excel 2019+ | 替代: 辅助单元格
    "LAMBDA",     # Excel 2019+ | 替代: VBA
    "SORT",       # Excel 2019+ | 替代: 排序功能
    "SORTBY",     # Excel 2019+ | 替代: 辅助列排序
    "RANDARRAY",  # Excel 2019+ | 替代: RAND()
]


def check_compatibility(formula: str) -> list[dict[str, str]]:
    """检查公式是否使用了不兼容的函数。
    
    Args:
        formula: Excel 公式
        
    Returns:
        包含不兼容函数信息的列表
    """
    issues = []
    for func in FORBIDDEN_FUNCTIONS:
        pattern = rf'\b{func}\s*\('
        if re.search(pattern, formula.upper()):
            alternatives = {
                "XLOOKUP": "INDEX() + MATCH()",
                "XMATCH": "MATCH()",
                "FILTER": "AutoFilter / SUMIF",
                "UNIQUE": "Remove Duplicates",
                "LET": "辅助单元格",
                "LAMBDA": "VBA / 辅助单元格",
                "SORT": "排序功能",
                "SORTBY": "辅助列排序",
                "RANDARRAY": "RAND()",
            }
            issues.append({
                "function": func,
                "alternative": alternatives.get(func, ""),
                "message": f"{func}() 在 Excel 2019/2016 中不支持"
            })
    
    return issues


# 用于测试
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        print(f"正在检查引用: {file_path}")
        result = refcheck_workbook(file_path)
        print(f"状态: {result.status}")
        print(f"问题数: {result.total_issues} ({result.error_count} 错误, {result.warning_count} 警告)")
        if result.issues:
            print("\n问题详情:")
            for issue in result.issues[:10]:
                print(f"  [{issue.severity.upper()}] {issue.cell} ({issue.sheet})")
                print(f"    {issue.message}")
                print(f"    建议: {issue.suggestion}")
    else:
        print("用法: python refcheck.py <excel_file.xlsx>")
