"""Excel 专业样式系统 - 提供财务、教育、科技等多种主题样式。

配色主题（参考 MiniMax styling.md）：
- grayscale: 通用，简约黑白，默认
- financial: 财务/金融，涨红跌绿，货币符号
- verdant: 教育/生态，绿色系
- dusk: 科技/创意，冷色调

样式规则：
- 隐藏网格线
- 内容从 B2 开始
- 蓝色字体 = 输入值/常数
- 黑色字体 = 公式结果
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@dataclass
class ColorPalette:
    """颜色调色板。"""
    # 主色
    primary: str = "000000"       # 主文本色
    # 背景
    background: str = "FFFFFF"    # 背景色
    alternate_bg: str = "F2F2F2"  # 交替行背景
    header_bg: str = "E6E6E6"    # 表头背景
    # 字体
    header_font: str = "000000"  # 表头字体色
    input_font: str = "1565C0"   # 输入值/常数 蓝色
    formula_font: str = "000000"  # 公式结果 黑色
    # 边框
    border: str = "CCCCCC"        # 边框色
    # 特殊色
    positive: str = "2E7D32"     # 正数/涨
    negative: str = "C62828"     # 负数/跌
    accent: str = "1976D2"       # 强调色


# 预定义主题
STYLE_THEMES = {
    "grayscale": ColorPalette(
        primary="000000",
        background="FFFFFF",
        alternate_bg="F5F5F5",
        header_bg="E0E0E0",
        header_font="000000",
        input_font="424242",
        formula_font="000000",
        border="BDBDBD",
        positive="2E7D32",
        negative="C62828",
        accent="616161",
    ),
    "financial": ColorPalette(
        primary="000000",
        background="FFFFFF",
        alternate_bg="FAFAFA",
        header_bg="1A237E",
        header_font="FFFFFF",
        input_font="1565C0",
        formula_font="000000",
        border="90CAF9",
        positive="2E7D32",
        negative="C62828",
        accent="FF6F00",
    ),
    "verdant": ColorPalette(
        primary="1B5E20",
        background="FFFFFF",
        alternate_bg="E8F5E9",
        header_bg="4CAF50",
        header_font="FFFFFF",
        input_font="1565C0",
        formula_font="1B5E20",
        border="A5D6A7",
        positive="2E7D32",
        negative="C62828",
        accent="388E3C",
    ),
    "dusk": ColorPalette(
        primary="263238",
        background="FFFFFF",
        alternate_bg="ECEFF1",
        header_bg="37474F",
        header_font="FFFFFF",
        input_font="0277BD",
        formula_font="263238",
        border="B0BEC5",
        positive="00838F",
        negative="D84315",
        accent="546E7A",
    ),
}


def _get_border(style: str = "thin") -> Border:
    """获取边框样式。"""
    side = Side(style=style, color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _get_header_border() -> Border:
    """获取表头边框样式。"""
    thin = Side(style="thin", color="CCCCCC")
    medium = Side(style="medium", color="000000")
    return Border(left=thin, right=thin, top=medium, bottom=medium)


def apply_professional_style(
    wb: Workbook,
    theme: str = "grayscale",
    hide_gridlines: bool = True,
    data_start_cell: str = "B2"
) -> Workbook:
    """为工作簿应用专业样式。
    
    Args:
        wb: openpyxl 工作簿对象
        theme: 主题名称（grayscale/financial/verdant/dusk）
        hide_gridlines: 是否隐藏网格线
        data_start_cell: 数据起始单元格
        
    Returns:
        应用样式后的工作簿
    """
    colors = STYLE_THEMES.get(theme, STYLE_THEMES["grayscale"])
    
    for ws in wb.worksheets:
        # 隐藏网格线
        if hide_gridlines:
            ws.sheet_view.showGridLines = False
        
        # 设置默认字体
        ws.font = Font(name="Calibri", size=11, color=colors.primary)
        
        # 应用标题行样式
        _apply_header_style(ws, colors)
        
        # 应用数据区域样式
        _apply_data_style(ws, colors, data_start_cell)
    
    return wb


def _apply_header_style(ws, colors: ColorPalette) -> None:
    """应用表头样式。"""
    # 表头背景填充
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=colors.header_bg
    )
    
    # 表头字体
    header_font = Font(
        name="Calibri",
        size=11,
        bold=True,
        color=colors.header_font
    )
    
    # 表头边框
    header_border = _get_header_border()
    
    # 表头对齐
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )
    
    # 遍历所有列应用表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment


def _apply_data_style(
    ws, 
    colors: ColorPalette,
    data_start_cell: str = "B2"
) -> None:
    """应用数据区域样式。"""
    # 交替行填充
    even_fill = PatternFill(
        fill_type="solid",
        fgColor=colors.background
    )
    odd_fill = PatternFill(
        fill_type="solid",
        fgColor=colors.alternate_bg
    )
    
    # 数据边框
    data_border = _get_border()
    
    # 数据对齐
    data_alignment = Alignment(
        horizontal="left",
        vertical="center"
    )
    
    # 数字对齐
    number_alignment = Alignment(
        horizontal="right",
        vertical="center"
    )
    
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    
    for row_idx in range(2, max_row + 1):  # 从第2行开始（跳过表头）
        fill = odd_fill if row_idx % 2 == 1 else even_fill
        
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = data_border
            cell.alignment = data_alignment
            
            # 数字单元格右对齐
            if cell.data_type in ('n',) or isinstance(cell.value, (int, float)):
                cell.alignment = number_alignment
                
                # 负数显示红色
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = Font(
                        name="Calibri",
                        size=11,
                        color=colors.negative
                    )


def apply_theme(
    wb: Workbook,
    theme: str,
    sheet_name: str | None = None
) -> Workbook:
    """为指定工作表应用主题。
    
    Args:
        wb: openpyxl 工作簿对象
        theme: 主题名称
        sheet_name: 工作表名称，None 表示所有工作表
        
    Returns:
        应用样式后的工作簿
    """
    if theme not in STYLE_THEMES:
        logger.warning(f"未知主题: {theme}，使用 grayscale")
        theme = "grayscale"
    
    colors = STYLE_THEMES[theme]
    
    target_sheets = [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets
    
    for ws in target_sheets:
        ws.sheet_view.showGridLines = False
        _apply_header_style(ws, colors)
        _apply_data_style(ws, colors)
    
    return wb


def set_column_widths(
    ws,
    widths: dict[str, float] | None = None,
    default_width: float = 12.0
) -> None:
    """设置列宽。
    
    Args:
        ws: 工作表对象
        widths: 列宽字典，键为列字母（如 'A'）或列范围（如 'A:C'），值为宽度
        default_width: 默认宽度
    """
    if widths:
        for col, width in widths.items():
            if ':' in col:
                # 列范围
                parts = col.split(':')
                start_col = parts[0]
                end_col = parts[1]
                from openpyxl.utils import column_index_from_string
                start_idx = column_index_from_string(start_col)
                end_idx = column_index_from_string(end_col)
                for idx in range(start_idx, end_idx + 1):
                    ws.column_dimensions[get_column_letter(idx)].width = width
            else:
                # 单列
                ws.column_dimensions[col].width = width
    else:
        # 自动调整所有列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大50
            ws.column_dimensions[column_letter].width = adjusted_width


def format_currency_column(
    ws,
    column: str,
    currency_symbol: str = "¥",
    decimal_places: int = 2
) -> None:
    """格式化货币列。
    
    Args:
        ws: 工作表对象
        column: 列字母（如 'B'）
        currency_symbol: 货币符号
        decimal_places: 小数位数
    """
    from openpyxl.utils import column_index_from_string
    
    col_idx = column_index_from_string(column)
    
    for cell in ws[column]:
        if isinstance(cell.value, (int, float)):
            format_str = f'{currency_symbol}#,##0.{"0" * decimal_places}'
            cell.number_format = format_str


def format_percentage_column(
    ws,
    column: str,
    decimal_places: int = 1
) -> None:
    """格式化百分比列。
    
    Args:
        ws: 工作表对象
        column: 列字母
        decimal_places: 小数位数
    """
    from openpyxl.utils import column_index_from_string
    
    col_idx = column_index_from_string(column)
    
    for cell in ws[column]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = f'0.{"0" * decimal_places}%'


def freeze_panes(
    ws,
    cell: str = "B2"
) -> None:
    """冻结窗格。
    
    Args:
        ws: 工作表对象
        cell: 冻结点单元格（如 'B2' 表示冻结首行首列）
    """
    ws.freeze_panes = cell


# 用于测试
if __name__ == "__main__":
    from openpyxl import Workbook
    
    # 创建测试工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试表"
    
    # 添加测试数据
    headers = ["姓名", "销售额", "增长率"]
    ws.append(headers)
    
    data = [
        ["张三", 15000, 0.15],
        ["李四", 23000, -0.05],
        ["王五", 18000, 0.08],
    ]
    for row in data:
        ws.append(row)
    
    # 应用样式
    print("应用 grayscale 主题...")
    apply_professional_style(wb, theme="grayscale")
    
    # 设置列宽
    set_column_widths(ws, {"A": 10, "B": 15, "C": 12})
    
    # 冻结首行
    freeze_panes(ws, "A2")
    
    # 保存
    output_path = "test_styled.xlsx"
    wb.save(output_path)
    print(f"已保存: {output_path}")
